from openpyxl import load_workbook
from database import get_connection

# Load Excel file
workbook = load_workbook("II CSE C.xlsx")
sheet = workbook.active

conn = get_connection()
cursor = conn.cursor()

count = 0

# Skip the header row
for row in sheet.iter_rows(min_row=2, values_only=True):

    register_number = str(row[1]).strip()
    student_name = str(row[2]).strip()

    cursor.execute("""
        INSERT IGNORE INTO students
        (register_number, student_name)
        VALUES (%s, %s)
    """, (register_number, student_name))

    count += 1

conn.commit()
conn.close()

print("--------------------------------")
print("Students Imported Successfully")
print("Total Students:", count)
print("--------------------------------")
